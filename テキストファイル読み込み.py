import numpy as np
import statistics
import pandas as pd
import openpyxl as excel

path = 'hukami.txt' #読み込みたいファイル名
A = []
with open(path) as f: #withはファイル操作で言う開く→読む→閉じるを一括でしてくれる。withの中身でopen→closeが自動でされる
    for s_line in f: #ファイルオブジェクトをそのままfor文で回すと、一行ずつ文字列として取得出来る。→https://note.nkmk.me/python-file-io-open-with/#moder
        A.append(s_line.split())
#print(A)

#wbname = "test.xlsx"
wb = excel.Workbook()
ws = wb.active
for j in range(len(A)):
    ws.cell(row=j+1, column=1, value=A[j][0])
    ws.cell(row=j+1, column=2, value=A[j][1])
    ws.cell(row=j+1, column=3, value=A[j][2])
    ws.cell(row=j+1, column=4, value=A[j][4])

# 保存する
wb.save("えくせる.xlsx")